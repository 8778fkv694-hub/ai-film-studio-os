from __future__ import annotations

import csv
import re
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path

import openpyxl


SOURCE_DIR = Path("/Users/yiliwen/Documents/全景插线理线(1)")
TIMING_XLSX = SOURCE_DIR / "时序对应表.xlsx"
OUT_ROOT = Path("/Users/yiliwen/开发/ai-film-studio-os/outputs/全景插线理线_切割重命名")

STAGE1_DIR = OUT_ROOT / "01_按时序表切割"
FINAL_VIDEO_DIR = OUT_ROOT / "02_按作业标准书命名" / "视频"
FINAL_IMAGE_DIR = OUT_ROOT / "02_按作业标准书命名" / "图片"
MANIFEST = OUT_ROOT / "映射表.csv"
LOG_FILE = OUT_ROOT / "处理日志.txt"

AVCONVERT = "/usr/bin/avconvert"
PRESET = "PresetHighestQuality"


@dataclass(frozen=True)
class WorkMap:
    sheet_no: int
    sheet: str
    step_no: str
    step_name: str
    cell: str
    confidence: str = "高"
    note: str = ""


MANUAL_MAP: dict[str, WorkMap] = {
    "1插线前主控板.jpg": WorkMap(2, "电源板与WiFi装配", "02", "固定主控板", "G20", "中", "原图为插线前主控板整体图"),
    "2插线后主控板.jpg": WorkMap(7, "走线", "03", "走线", "O8", "中", "时序表作为走线扎带视频附图"),
    "3增压泵走线.jpg": WorkMap(1, "增压泵线走线", "04", "增压泵线走线", "O20"),
    "4适配器连接线.jpg": WorkMap(2, "电源板与WiFi装配", "03", "固定适配器连接线", "K21"),
    "5WIFI打螺丝走线.jpg": WorkMap(2, "电源板与WiFi装配", "04", "固定WiFi模块组件", "O20"),
    "6适配器输出端子.jpg": WorkMap(3, "插线1", "01", "插线适配器输出端子", "G7"),
    "7适配器输入端子.jpg": WorkMap(3, "插线1", "02", "插线适配器输入端子", "M7"),
    "8WIFI模块组件线.jpg": WorkMap(3, "插线1", "03", "插线WIFI模块组件", "G20"),
    "9加热体线.jpg": WorkMap(3, "插线1", "04", "插线加热体", "L20"),
    "10净水阀线.jpg": WorkMap(4, "插线2", "01", "插线净水阀", "G7"),
    "11废水阀线.jpg": WorkMap(4, "插线2", "02", "插线废水阀", "K7"),
    "12流量计线.jpg": WorkMap(4, "插线2", "03", "插线流量计", "O7"),
    "13净水阀阀1粉线.jpg": WorkMap(4, "插线2", "04", "插线净水阀阀1", "G20"),
    "14热水阀线.jpg": WorkMap(4, "插线2", "05", "插线热水泵", "K20"),
    "15温度传感器线.jpg": WorkMap(4, "插线2", "06", "插线温度传感器", "O20"),
    "16温度1.jpg": WorkMap(5, "插线3", "01", "插线温度1", "G7"),
    "17显示板连接线.jpg": WorkMap(5, "插线3", "02", "插线显示板连接线", "K7"),
    "18高压开关线.jpg": WorkMap(5, "插线3", "03", "插线高压开关", "O7"),
    "19进水电磁阀线.jpg": WorkMap(5, "插线3", "04", "插线进水电磁阀", "G20"),
    "20进水电动阀.jpg": WorkMap(5, "插线3", "05", "插线进水电动阀", "K20"),
    "21动紫外.jpg": WorkMap(5, "插线3", "06", "插线动紫外", "O20"),
    "22阀2黑.jpg": WorkMap(6, "插线4", "01", "插线阀2", "G7"),
    "23增压泵连接线.jpg": WorkMap(6, "插线4", "02", "插线增压泵", "K7"),
    "24显示板走线.jpg": WorkMap(7, "走线", "01", "显示板走线", "H7"),
    "25显示板走线2.jpg": WorkMap(7, "走线", "02", "显示板走线", "K7"),
    "26扎线1.jpg": WorkMap(7, "走线", "03", "走线扎线1", "O8", "中", "作业标准书该组为走线/扎线连续图"),
    "27扎线2.jpg": WorkMap(7, "走线", "04", "走线扎线2", "G20", "中", "作业标准书该组为走线/扎线连续图"),
    "28扎线3.jpg": WorkMap(7, "走线", "05", "走线扎线3", "K20", "中", "作业标准书该组为走线/扎线连续图"),
    "29扎线4.jpg": WorkMap(7, "走线", "06", "走线扎线4", "O20", "中", "作业标准书该组为走线/扎线连续图"),
    "30扎线5.jpg": WorkMap(7, "走线", "07", "走线扎线5", "G20", "中", "时序表作为该视频起始扎线图"),
    "31扎线6（供应商来料扎好）.jpg": WorkMap(7, "走线", "08", "走线扎线6供应商来料扎好", "K20", "中", "时序表作为该视频附图"),
    "32信号线穿线.jpg": WorkMap(8, "固定信号线", "01A", "信号线穿线", "G7", "中", "时序表细分为穿线与固定"),
    "32信号线固定.jpg": WorkMap(8, "固定信号线", "01B", "固定信号线卡", "G7"),
    "33电源线插入.jpg": WorkMap(11, "走线2", "03", "电源线插入", "O7", "中", "时序表细分电源线安装动作"),
    "34电源线固定.jpg": WorkMap(9, "固定电源线固定", "02", "固定电源线", "K7"),
    "35电源线从PE管下方穿过.jpg": WorkMap(11, "走线2", "02", "电源线从PE管下方穿过", "K7"),
    "36电源线从加热体线下方穿过.jpg": WorkMap(11, "走线2", "01", "电源线从加热体线下方穿过", "H7"),
    "37增压泵线走线放置水路板后侧.jpg": WorkMap(1, "增压泵线走线", "04B", "增压泵线走线放置水路板后侧", "O20"),
    "38接地线固定.jpg": WorkMap(10, "固定接地线", "05", "固定接地线", "O20"),
}


def sanitize(value: str) -> str:
    value = value.strip()
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", "", value)
    value = value.replace("__", "_")
    return value[:120].strip("._ ")


def parse_time(value: object) -> tuple[str, float | None, float | None]:
    text = "" if value is None else str(value).strip()
    if not text:
        return "", None, None
    if text == "整段":
        return text, None, None
    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)\s*s?", text)
    if not match:
        raise ValueError(f"Unsupported time value: {text}")
    start = float(match.group(1))
    end = float(match.group(2))
    if end <= start:
        raise ValueError(f"Invalid time range: {text}")
    return text, start, end - start


def load_timing_rows() -> tuple[list[dict[str, object]], dict[str, dict[str, object]]]:
    wb = openpyxl.load_workbook(TIMING_XLSX, data_only=True)
    ws = wb.active
    rows: list[dict[str, object]] = []
    images: dict[str, dict[str, object]] = {}
    current_video = None
    current_video_label = None
    order = 0

    for excel_row in range(2, ws.max_row + 1):
        video_label = ws.cell(excel_row, 1).value
        video_file = ws.cell(excel_row, 2).value
        if video_file:
            current_video = str(video_file).strip()
            current_video_label = str(video_label).strip() if video_label else ""

        image_file = ws.cell(excel_row, 3).value
        time_text = ws.cell(excel_row, 4).value
        if image_file:
            order += 1
            time_label, start, duration = parse_time(time_text)
            record = {
                "order": order,
                "excel_row": excel_row,
                "video_label": current_video_label,
                "video_file": current_video,
                "image_file": str(image_file).strip(),
                "time_label": time_label,
                "start": start,
                "duration": duration,
            }
            rows.append(record)
            images[record["image_file"]] = record

        for col in range(5, ws.max_column + 1):
            extra = ws.cell(excel_row, col).value
            if extra:
                extra_name = str(extra).strip()
                images.setdefault(
                    extra_name,
                    {
                        "order": None,
                        "excel_row": excel_row,
                        "video_label": current_video_label,
                        "video_file": current_video,
                        "image_file": extra_name,
                        "time_label": "",
                        "start": None,
                        "duration": None,
                    },
                )
    return rows, images


def final_base(image_file: str) -> str:
    stem = Path(image_file).stem
    work = MANUAL_MAP.get(image_file)
    if work is None:
        return sanitize(stem)
    return sanitize(
        f"{work.sheet_no:02d}_{work.sheet}_步骤{work.step_no}_{work.step_name}_{work.cell}_{stem}"
    )


def cut_or_copy_video(record: dict[str, object], stage1_path: Path) -> None:
    video_file = record["video_file"]
    if not video_file:
        raise ValueError(f"Missing video for {record['image_file']}")
    source = SOURCE_DIR / str(video_file)
    if not source.exists():
        raise FileNotFoundError(source)

    if record["time_label"] == "整段":
        shutil.copy2(source, stage1_path)
        return

    start = record["start"]
    duration = record["duration"]
    if start is None or duration is None:
        return

    cmd = [
        AVCONVERT,
        "--source",
        str(source),
        "--preset",
        PRESET,
        "--start",
        f"{start:g}",
        "--duration",
        f"{duration:g}",
        "--output",
        str(stage1_path),
        "--replace",
    ]
    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)


def main() -> None:
    for path in [STAGE1_DIR, FINAL_VIDEO_DIR, FINAL_IMAGE_DIR]:
        path.mkdir(parents=True, exist_ok=True)

    timing_rows, image_records = load_timing_rows()
    log_lines: list[str] = []
    manifest_rows: list[dict[str, object]] = []

    for record in timing_rows:
        if not record["time_label"]:
            continue
        base = final_base(str(record["image_file"]))
        src_ext = Path(str(record["video_file"])).suffix.lower() if record["video_file"] else ".mp4"
        stage_ext = src_ext if record["time_label"] == "整段" else ".m4v"
        stage1_name = sanitize(
            f"{int(record['order']):02d}_{Path(str(record['image_file'])).stem}_{Path(str(record['video_file'])).stem}_{record['time_label']}"
        ) + stage_ext
        stage1_path = STAGE1_DIR / stage1_name
        cut_or_copy_video(record, stage1_path)

        final_video = FINAL_VIDEO_DIR / (base + stage_ext)
        shutil.copy2(stage1_path, final_video)
        record["stage1_video"] = str(stage1_path)
        record["final_video"] = str(final_video)
        log_lines.append(f"video: {record['image_file']} -> {final_video.name}")

    source_images = sorted(SOURCE_DIR.glob("*.jpg"), key=lambda p: p.name)
    for image_path in source_images:
        image_file = image_path.name
        base = final_base(image_file)
        final_image = FINAL_IMAGE_DIR / (base + image_path.suffix.lower())
        shutil.copy2(image_path, final_image)

        timing = image_records.get(image_file, {})
        work = MANUAL_MAP.get(image_file)
        final_video = ""
        if isinstance(timing, dict):
            final_video = str(timing.get("final_video", ""))

        manifest_rows.append(
            {
                "原图片": image_file,
                "原视频": timing.get("video_file", "") if isinstance(timing, dict) else "",
                "时间段": timing.get("time_label", "") if isinstance(timing, dict) else "",
                "一阶段切割视频": timing.get("stage1_video", "") if isinstance(timing, dict) else "",
                "最终视频": final_video,
                "最终图片": str(final_image),
                "作业标准书Sheet": work.sheet if work else "",
                "Sheet序号": f"{work.sheet_no:02d}" if work else "",
                "步骤": f"步骤{work.step_no}" if work else "",
                "步骤名称": work.step_name if work else "",
                "工作格/图片格": work.cell if work else "",
                "置信度": work.confidence if work else "",
                "备注": work.note if work else "",
            }
        )
        log_lines.append(f"image: {image_file} -> {final_image.name}")

    with MANIFEST.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(manifest_rows[0].keys()))
        writer.writeheader()
        writer.writerows(manifest_rows)

    LOG_FILE.write_text("\n".join(log_lines) + "\n", encoding="utf-8")
    print(f"cut videos: {sum(1 for r in timing_rows if r['time_label'])}")
    print(f"renamed images: {len(source_images)}")
    print(f"manifest: {MANIFEST}")
    print(f"output: {OUT_ROOT}")


if __name__ == "__main__":
    main()
