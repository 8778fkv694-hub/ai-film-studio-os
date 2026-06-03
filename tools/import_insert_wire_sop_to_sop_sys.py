from __future__ import annotations

import csv
import hashlib
import mimetypes
import os
import re
import shutil
import sqlite3
import subprocess
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import openpyxl


SOP_SYS = Path("/Users/yiliwen/开发/sop-sys")
DB_PATH = SOP_SYS / "database" / "dev.sqlite"
RUNTIME_ASSETS = SOP_SYS / "runtime" / "data" / "current" / "assets"
DOCS_DIR = SOP_SYS / "docs"

SOURCE_XLSX = Path("/Users/yiliwen/Documents/插线理线--YCZ-JT1600-02-X20(G)作业标准书.xlsx")
SOURCE_PDF = Path("/Users/yiliwen/Documents/插线理线--YCZ-JT1600-02-X20(G)作业标准书.pdf")
DECOMPOSED_ROOT = Path("/Users/yiliwen/开发/ai-film-studio-os/outputs/全景插线理线_切割重命名")
MANIFEST = DECOMPOSED_ROOT / "映射表.csv"
FINAL_IMAGE_DIR = DECOMPOSED_ROOT / "02_按作业标准书命名" / "图片"
FINAL_VIDEO_DIR = DECOMPOSED_ROOT / "02_按作业标准书命名" / "视频"
STILL_MISSING = DECOMPOSED_ROOT / "03_从作业标准书补提图片" / "仍缺清单.csv"

TMP_PDF_RENDER = Path("/private/tmp/sop_sys_insert_wire_pdf_pages")
PDFTOPPM = Path("/opt/homebrew/bin/pdftoppm")

ASSET_SET = "ycz-jt1600-02-x20g"
MODEL_NAME = "YCZ-JT1600-02-X20(G)"
SERIES_NAME = "YCZ-JT1600-02-X20(G) 插线理线"
PRODUCT_NAME = "净水机"
PART_NAME = "插线理线"
FORM_NO = "FT[厨电]OTD-ZB100-01B"
REVISION = "A0"
TAG = "v1.0"

VIDEO_AUDIT_BLOCKLIST = {
    "1插线前主控板.jpg",
    "25显示板走线2.jpg",
    "26扎线1.jpg",
    "27扎线2.jpg",
    "28扎线3.jpg",
    "29扎线4.jpg",
    "30扎线5.jpg",
    "32信号线穿线.jpg",
    "33电源线插入.jpg",
}

AUDITED_ASSET_RENAMES = {
    "33电源线插入.jpg": {
        "最终图片": "09_固定电源线固定_步骤01_固定电源线_G7_33电源线插入.jpg",
        "作业标准书Sheet": "固定电源线固定",
        "Sheet序号": "09",
        "步骤": "步骤01",
        "步骤名称": "固定电源线",
        "工作格/图片格": "G7",
        "置信度": "高",
        "备注": "内容核对后从走线2步骤03移至固定电源线步骤01；同段视频不够确定，未挂载。",
    },
    "37增压泵线走线放置水路板后侧.jpg": {
        "最终图片": "11_走线2_步骤03_增压泵线束磁环放置于水路板后侧_O7_37增压泵线走线放置水路板后侧.jpg",
        "最终视频": "11_走线2_步骤03_增压泵线束磁环放置于水路板后侧_O7_37增压泵线走线放置水路板后侧.m4v",
        "作业标准书Sheet": "走线2",
        "Sheet序号": "11",
        "步骤": "步骤03",
        "步骤名称": "增压泵线束磁环放置于水路板后侧",
        "工作格/图片格": "O7",
        "置信度": "高",
        "备注": "内容核对后按作业标准书走线2步骤03文字移动；原放在增压泵线走线步骤04B。",
    },
}


@dataclass(frozen=True)
class ImportedAsset:
    src: str
    target_path: Path


def require_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(path)


def clean_segment(value: str) -> str:
    value = str(value).strip()
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("._ ")[:150] or "asset"


def slug_for_id(value: str) -> str:
    value = str(value).strip().lower()
    value = re.sub(r"[^a-z0-9\u4e00-\u9fa5]+", "-", value)
    return value.strip("-") or "sop"


def runtime_url(relative_path: str) -> str:
    parts = [quote(part) for part in relative_path.split("/") if part]
    return "/api/files/runtime-assets/current/" + "/".join(parts)


def copy_asset(source: Path, category: str, preferred_name: str, force_ext: str | None = None) -> ImportedAsset:
    require_file(source)
    extension = force_ext if force_ext else source.suffix.lower()
    file_name = clean_segment(Path(preferred_name).stem) + extension
    relative_path = f"{category}/{ASSET_SET}/{file_name}"
    target_path = RUNTIME_ASSETS / category / ASSET_SET / file_name
    target_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target_path)
    return ImportedAsset(src=runtime_url(relative_path), target_path=target_path)


def render_pdf_pages() -> list[Path]:
    require_file(SOURCE_PDF)
    if TMP_PDF_RENDER.exists():
        shutil.rmtree(TMP_PDF_RENDER)
    TMP_PDF_RENDER.mkdir(parents=True, exist_ok=True)

    subprocess.run(
        [
            str(PDFTOPPM),
            "-png",
            "-r",
            "120",
            str(SOURCE_PDF),
            str(TMP_PDF_RENDER / "page"),
        ],
        check=True,
    )
    pages = sorted(TMP_PDF_RENDER.glob("page-*.png"))
    if len(pages) != 11:
        raise RuntimeError(f"Expected 11 rendered PDF pages, got {len(pages)}")
    return pages


def parse_step_token(value: str) -> tuple[int, str, str]:
    text = str(value).strip()
    match = re.search(r"(\d+)([A-Za-z]?)", text)
    if not match:
        return 999, "", text
    return int(match.group(1)), match.group(2).upper(), text


def parse_workbook_metadata() -> tuple[list[str], dict[tuple[str, int], str], dict[tuple[str, int], str]]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    sheet_names = [ws.title.strip() for ws in wb.worksheets]
    flow_titles: dict[tuple[str, int], str] = {}
    requirement_text: dict[tuple[str, int], str] = {}

    for ws in wb.worksheets:
        sheet = ws.title.strip()
        for row in [7, 10, 13, 16, 19, 22]:
            seq = ws.cell(row, 2).value
            title = ws.cell(row, 3).value
            if isinstance(seq, int) and title:
                flow_titles[(sheet, seq)] = str(title).strip()

        for label_row, req_row in [(6, 15), (19, 28)]:
            for col in [7, 11, 15]:
                label = ws.cell(label_row, col).value
                if not isinstance(label, str):
                    continue
                match = re.search(r"步骤\s*(\d+)", label)
                if not match:
                    continue
                req = ws.cell(req_row, col).value
                if req:
                    requirement_text[(sheet, int(match.group(1)))] = str(req).strip()

    return sheet_names, flow_titles, requirement_text


def read_manifest_rows() -> list[dict[str, str]]:
    require_file(MANIFEST)
    with MANIFEST.open("r", encoding="utf-8-sig", newline="") as f:
        rows = [row for row in csv.DictReader(f)]
    apply_media_audit_overrides(rows)
    return rows


def append_note(row: dict[str, str], note: str) -> None:
    existing = row.get("备注", "").strip()
    if note in existing:
        return
    row["备注"] = f"{existing}；{note}" if existing else note


def ensure_asset_copy(source_text: str, target_dir: Path, target_name: str) -> str:
    if not source_text:
        return ""
    source = Path(source_text)
    target = target_dir / target_name
    if source.resolve() != target.resolve():
        require_file(source)
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)
    return str(target)


def apply_media_audit_overrides(rows: list[dict[str, str]]) -> None:
    """Apply visual/semantic audit corrections before building SOP records.

    The time table contains useful cuts, but several rows are continuous helper
    videos or were attached to a similarly named but wrong SOP step. Uncertain
    videos are intentionally left unattached.
    """
    for row in rows:
        original_image = row.get("原图片", "").strip()

        if original_image in AUDITED_ASSET_RENAMES:
            override = AUDITED_ASSET_RENAMES[original_image]
            if override.get("最终图片"):
                row["最终图片"] = ensure_asset_copy(row.get("最终图片", ""), FINAL_IMAGE_DIR, override["最终图片"])
            if override.get("最终视频"):
                row["最终视频"] = ensure_asset_copy(row.get("最终视频", ""), FINAL_VIDEO_DIR, override["最终视频"])
            for key, value in override.items():
                if key in {"最终图片", "最终视频"}:
                    continue
                row[key] = value

        if original_image in VIDEO_AUDIT_BLOCKLIST:
            row["最终视频"] = ""
            append_note(row, "内容核对后视频不够确定，电子版不挂载该视频")


def read_still_missing_rows() -> list[dict[str, str]]:
    if not STILL_MISSING.exists():
        return []
    with STILL_MISSING.open("r", encoding="utf-8-sig", newline="") as f:
        return [row for row in csv.DictReader(f)]


def build_media_assets() -> tuple[dict[str, ImportedAsset], dict[str, ImportedAsset], list[ImportedAsset]]:
    image_assets: dict[str, ImportedAsset] = {}
    video_assets: dict[str, ImportedAsset] = {}

    for source in sorted(FINAL_IMAGE_DIR.glob("*")):
        if source.name.startswith(".") or source.suffix.lower() not in {".jpg", ".jpeg", ".png", ".gif", ".webp"}:
            continue
        image_assets[str(source)] = copy_asset(source, "images", source.name)

    for source in sorted(FINAL_VIDEO_DIR.glob("*")):
        if source.name.startswith(".") or source.suffix.lower() not in {".mp4", ".m4v", ".mov", ".webm"}:
            continue
        # The backend only advertises video/mp4 for .mp4. M4V files are ISO media,
        # so keep the bytes and expose them with .mp4 extension for browser playback.
        video_assets[str(source)] = copy_asset(source, "videos", source.name, force_ext=".mp4")

    paper_assets: list[ImportedAsset] = []
    for index, source in enumerate(render_pdf_pages(), 1):
        paper_assets.append(copy_asset(source, "images", f"{ASSET_SET}_paper_page_{index:02d}.png"))

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_PDF, DOCS_DIR / SOURCE_PDF.name)

    return image_assets, video_assets, paper_assets


def group_manifest_steps(
    manifest_rows: list[dict[str, str]],
    image_assets: dict[str, ImportedAsset],
    video_assets: dict[str, ImportedAsset],
    requirement_text: dict[tuple[str, int], str],
) -> dict[str, list[dict[str, object]]]:
    by_sheet: dict[str, list[dict[str, object]]] = {}

    for row in manifest_rows:
        sheet = row.get("作业标准书Sheet", "").strip()
        if not sheet:
            continue
        step_label = row.get("步骤", "").strip().replace("步骤", "")
        step_number, step_suffix, raw_token = parse_step_token(step_label)
        step_name = row.get("步骤名称", "").strip() or Path(row.get("原图片", "")).stem or raw_token
        cell = row.get("工作格/图片格", "").strip()
        original_image = row.get("原图片", "").strip()
        image_path = row.get("最终图片", "").strip()
        video_path = row.get("最终视频", "").strip()
        image_asset = image_assets.get(image_path)
        video_asset = video_assets.get(video_path)

        media = []
        if image_asset:
            media.append(
                {
                    "id": f"image-{len(media) + 1}",
                    "kind": "image",
                    "src": image_asset.src,
                    "title": f"{step_name} 图片（{cell}）" if cell else f"{step_name} 图片",
                }
            )
        if video_asset:
            media.append(
                {
                    "id": f"video-{len(media) + 1}",
                    "kind": "video",
                    "src": video_asset.src,
                    "title": f"{step_name} 视频（{row.get('时间段', '').strip()}）",
                    "playSequential": False,
                    "playInterval": 0.6,
                    "playLoop": False,
                }
            )

        detail = requirement_text.get((sheet, step_number), "")
        note = row.get("备注", "").strip()
        if note:
            detail = f"{detail}\n备注：{note}".strip()
        if not detail:
            detail = f"按作业标准书 {sheet} {row.get('步骤', '').strip()} 执行。"

        by_sheet.setdefault(sheet, []).append(
            {
                "stepNumber": step_number,
                "stepSuffix": step_suffix,
                "rawToken": raw_token,
                "name": step_name,
                "cell": cell,
                "originalImage": original_image,
                "detail": detail,
                "media": media,
                "hasVideo": bool(video_asset),
            }
        )

    return by_sheet


def add_missing_standard_steps(
    by_sheet: dict[str, list[dict[str, object]]],
    missing_rows: list[dict[str, str]],
    requirement_text: dict[tuple[str, int], str],
) -> None:
    for row in missing_rows:
        position = row.get("作业标准书位置", "")
        if "_" not in position:
            continue
        sheet = position.split("_", 1)[1].strip()
        step_label = row.get("步骤", "").replace("步骤", "").strip()
        step_number, step_suffix, raw_token = parse_step_token(step_label)
        name = row.get("内容", "").strip() or f"步骤{raw_token}"
        detail = requirement_text.get((sheet, step_number), "") or "原作业标准书有此步骤，但源图片、解构视频和 Excel 嵌入图中均未找到可补素材。"
        by_sheet.setdefault(sheet, []).append(
            {
                "stepNumber": step_number,
                "stepSuffix": step_suffix,
                "rawToken": raw_token,
                "name": name,
                "cell": "",
                "originalImage": "",
                "detail": detail,
                "media": [],
                "hasVideo": False,
            }
        )


def add_flow_steps_without_media(
    by_sheet: dict[str, list[dict[str, object]]],
    flow_titles: dict[tuple[str, int], str],
    requirement_text: dict[tuple[str, int], str],
) -> None:
    for (sheet, step_number), title in sorted(flow_titles.items(), key=lambda item: (item[0][0], item[0][1])):
        existing = by_sheet.get(sheet, [])
        if any(int(item["stepNumber"]) == step_number for item in existing):
            continue

        detail = requirement_text.get((sheet, step_number), "")
        if not detail:
            detail = "作业标准书流程步骤，当前源图片、解构视频和 Excel 嵌入图中均未找到可补素材。"

        by_sheet.setdefault(sheet, []).append(
            {
                "stepNumber": step_number,
                "stepSuffix": "",
                "rawToken": f"{step_number:02d}",
                "name": title,
                "cell": "",
                "originalImage": "",
                "detail": detail,
                "media": [],
                "hasVideo": False,
            }
        )


def make_sop_records(
    sheet_names: list[str],
    by_sheet: dict[str, list[dict[str, object]]],
    flow_titles: dict[tuple[str, int], str],
    paper_assets: list[ImportedAsset],
) -> list[dict[str, object]]:
    records = []
    updated_at = datetime.now().replace(microsecond=0).isoformat()

    for sheet_index, sheet in enumerate(sheet_names, 1):
        document_id = f"ycz-jt1600-02-x20g-{sheet_index:02d}-{slug_for_id(sheet)}"
        rows = by_sheet.get(sheet, [])
        rows.sort(key=lambda item: (int(item["stepNumber"]), str(item["stepSuffix"]), str(item["name"])))

        # Remove exact duplicates caused by image/video manifest updates while
        # preserving distinct cells or suffixes.
        seen = set()
        deduped = []
        for item in rows:
            key = (item["stepNumber"], item["stepSuffix"], item["name"], item["cell"], item["originalImage"])
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)

        steps = []
        for step_order, item in enumerate(deduped, 1):
            step_number = int(item["stepNumber"])
            suffix = str(item["stepSuffix"])
            raw_token = str(item["rawToken"])
            source_bits = [f"{sheet} / 步骤{raw_token}"]
            if item["cell"]:
                source_bits.append(f"工作格 {item['cell']}")
            if item["originalImage"]:
                source_bits.append(str(item["originalImage"]))
            title = f"步骤{raw_token} {item['name']}"
            if suffix and not raw_token.endswith(suffix):
                title = f"步骤{step_number:02d}{suffix} {item['name']}"
            steps.append(
                {
                    "id": f"step-{step_order:02d}",
                    "order": step_order,
                    "title": title,
                    "action": str(item["name"]),
                    "detail": str(item["detail"]),
                    "sourceLabel": " / ".join(source_bits),
                    "controlPoint": flow_titles.get((sheet, step_number), ""),
                    "media": item["media"],
                }
            )

        paper = paper_assets[sheet_index - 1]
        records.append(
            {
                "id": document_id,
                "sopNo": f"SOP-YCZ-JT1600-02-X20G-{sheet_index:02d}",
                "seriesName": SERIES_NAME,
                "productName": PRODUCT_NAME,
                "modelName": MODEL_NAME,
                "partName": PART_NAME,
                "title": sheet,
                "formNo": FORM_NO,
                "processCode": f"X20G-{sheet_index:02d}",
                "deviceName": "无",
                "moldNo": "无",
                "department": "净水器组装",
                "revision": REVISION,
                "tag": TAG,
                "pageNumber": sheet_index,
                "status": "published",
                "updatedAt": updated_at,
                "summary": f"{sheet}，电子版 {len(steps)} 个步骤，纸版第 {sheet_index} 页。",
                "validUntil": None,
                "paperPages": [
                    {
                        "pageNumber": sheet_index,
                        "image": paper.src,
                        "caption": f"{sheet} 纸版 PDF 第 {sheet_index} 页",
                        "note": f"来源：{SOURCE_PDF.name}",
                    }
                ],
                "steps": steps,
                "materials": [],
                "tools": [],
                "qualityRisks": [],
                "selfCheckFrequency": "每件",
            }
        )

    return records


def backup_database() -> Path:
    require_file(DB_PATH)
    backup = DB_PATH.with_name(f"{DB_PATH.name}.bak-{datetime.now().strftime('%Y%m%d-%H%M%S')}")
    shutil.copy2(DB_PATH, backup)
    return backup


def insert_records(records: list[dict[str, object]]) -> None:
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA foreign_keys = ON")
    try:
        with db:
            ids = [record["id"] for record in records]
            for document_id in ids:
                db.execute("DELETE FROM sop_documents WHERE id = ?", (document_id,))

            for record in records:
                db.execute(
                    """
                    INSERT INTO sop_documents (
                      id, sop_no, series_id, series_name, product_name, model_name,
                      part_name, title, form_no, process_code, device_name, mold_no,
                      department, valid_until, status, deleted_at, deleted_note, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, ?)
                    """,
                    (
                        record["id"],
                        record["sopNo"],
                        slug_for_id(record["seriesName"]),
                        record["seriesName"],
                        record["productName"],
                        record["modelName"],
                        record["partName"],
                        record["title"],
                        record["formNo"],
                        record["processCode"],
                        record["deviceName"],
                        record["moldNo"],
                        record["department"],
                        record["validUntil"],
                        record["status"],
                        record["updatedAt"],
                    ),
                )

                revision_id = f"{record['id']}@rev-{record['revision']}"
                db.execute(
                    """
                    INSERT INTO sop_revisions (
                      id, document_id, revision_code, page_number, status, is_current,
                      is_published, updated_at, published_at, summary, tag,
                      controlled_statement, source_note, self_check_frequency
                    ) VALUES (?, ?, ?, ?, ?, 1, 1, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        revision_id,
                        record["id"],
                        record["revision"],
                        record["pageNumber"],
                        record["status"],
                        record["updatedAt"],
                        record["updatedAt"],
                        record["summary"],
                        record["tag"],
                        "纸版受控文件为唯一执行依据，电子版用于现场展示。",
                        f"电子版素材由 {DECOMPOSED_ROOT} 解构生成；纸版来自 {SOURCE_PDF.name}",
                        record["selfCheckFrequency"],
                    ),
                )

                for index, page in enumerate(record["paperPages"], 1):
                    db.execute(
                        """
                        INSERT INTO sop_pages (
                          id, revision_id, page_number, image_src, caption, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            f"{revision_id}:page-{index}",
                            revision_id,
                            page["pageNumber"],
                            page["image"],
                            page["caption"],
                            page["note"],
                        ),
                    )

                for step in record["steps"]:
                    step_id = f"{revision_id}:{step['id']}"
                    db.execute(
                        """
                        INSERT INTO sop_steps (
                          id, revision_id, step_order, title, action_text,
                          detail_text, source_label, control_point
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            step_id,
                            revision_id,
                            step["order"],
                            step["title"],
                            step["action"],
                            step["detail"],
                            step["sourceLabel"],
                            step.get("controlPoint") or None,
                        ),
                    )
                    for media_index, media in enumerate(step["media"], 1):
                        db.execute(
                            """
                            INSERT INTO step_media (
                              id, step_id, media_order, media_kind, src, title,
                              poster, crop_x, crop_y, crop_width, crop_height,
                              is_default, annotations_json, play_sequential,
                              play_interval, play_loop
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, NULL, NULL, NULL, NULL, 1, NULL, ?, ?, ?)
                            """,
                            (
                                f"{step_id}:{media['id']}",
                                step_id,
                                media_index,
                                media["kind"],
                                media["src"],
                                media["title"],
                                media.get("poster"),
                                1 if media.get("playSequential") else 0,
                                media.get("playInterval", 0.6),
                                1 if media.get("playLoop") else 0,
                            ),
                        )

                db.execute(
                    """
                    INSERT INTO audit_logs (id, action, target_type, target_id, payload_json)
                    VALUES (?, 'sop.import.create', 'sop_document', ?, ?)
                    """,
                    (
                        hashlib.sha1(f"{record['id']}:{record['updatedAt']}".encode("utf-8")).hexdigest(),
                        record["id"],
                        f'{{"note":"导入插线理线 SOP，含纸版 PDF 页图和电子版图片/视频素材","stepCount":{len(record["steps"])}}}',
                    ),
                )
    finally:
        db.close()


def write_import_summary(records: list[dict[str, object]], backup_path: Path) -> Path:
    summary_path = SOP_SYS / "runtime" / "insert-wire-sop-import-summary.csv"
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with summary_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["SOP ID", "SOP编号", "标题", "纸版页", "电子步骤数", "媒体数", "数据库备份"])
        for record in records:
            media_count = sum(len(step["media"]) for step in record["steps"])
            writer.writerow(
                [
                    record["id"],
                    record["sopNo"],
                    record["title"],
                    record["pageNumber"],
                    len(record["steps"]),
                    media_count,
                    str(backup_path),
                ]
            )
    return summary_path


def main() -> None:
    for path in [SOURCE_XLSX, SOURCE_PDF, MANIFEST, FINAL_IMAGE_DIR, FINAL_VIDEO_DIR, DB_PATH]:
        require_file(path)

    sheet_names, flow_titles, requirement_text = parse_workbook_metadata()
    manifest_rows = read_manifest_rows()
    still_missing_rows = read_still_missing_rows()
    image_assets, video_assets, paper_assets = build_media_assets()
    by_sheet = group_manifest_steps(manifest_rows, image_assets, video_assets, requirement_text)
    add_missing_standard_steps(by_sheet, still_missing_rows, requirement_text)
    add_flow_steps_without_media(by_sheet, flow_titles, requirement_text)
    records = make_sop_records(sheet_names, by_sheet, flow_titles, paper_assets)
    backup = backup_database()
    insert_records(records)
    summary = write_import_summary(records, backup)

    print(f"imported_sops={len(records)}")
    print(f"images={len(image_assets)} videos={len(video_assets)} paper_pages={len(paper_assets)}")
    print(f"backup={backup}")
    print(f"summary={summary}")
    print(f"docs_pdf={DOCS_DIR / SOURCE_PDF.name}")


if __name__ == "__main__":
    main()
